"""Build the Excel workbook deliverable for the portfolio project.

Run from the project root:
    python scripts/build_excel_workbook.py
"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_PATH = ROOT / "data" / "processed" / "food_delivery_cleaned.csv"
OUTPUT_PATH = ROOT / "excel" / "food_delivery_profitability_workbook.xlsx"

BLUE = "1F4E78"
LIGHT_BLUE = "D9EAF7"
GREEN = "E2F0D9"
AMBER = "FFF2CC"


def write_table(ws, dataframe, start_row=1, start_col=1):
    for col_idx, column in enumerate(dataframe.columns, start_col):
        cell = ws.cell(start_row, col_idx, column)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center")
    for row_idx, row in enumerate(dataframe.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, start_col):
            ws.cell(row_idx, col_idx, value)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = ws.cell(start_row + 1, start_col).coordinate
    for col in range(start_col, start_col + len(dataframe.columns)):
        ws.column_dimensions[get_column_letter(col)].width = min(24, max(12, len(str(ws.cell(start_row, col).value)) + 4))


def style_title(ws, title):
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color=BLUE)
    ws.merge_cells("A1:H1")


def main():
    df = pd.read_csv(DATA_PATH, parse_dates=["order_datetime"])

    kpis = pd.DataFrame(
        [
            ["Total Orders", len(df)],
            ["Gross Profit", round(df["gross_profit"].sum(), 2)],
            ["Average Order Value", round(df["order_value"].mean(), 2)],
            ["Average Margin %", round(df["profit_margin_pct"].mean(), 2)],
            ["Repeat Rate %", round(df["repeat_customer_flag"].mean() * 100, 2)],
            ["Delay Rate %", round(df["delivery_delay_flag"].mean() * 100, 2)],
            ["Average Rating", round(df["rating"].mean(), 2)],
            ["Total Discount", round(df["discount_amount"].sum(), 2)],
        ],
        columns=["KPI", "Value"],
    )

    city = (
        df.groupby("city")
        .agg(
            orders=("order_id", "count"),
            gross_profit=("gross_profit", "sum"),
            avg_margin_pct=("profit_margin_pct", "mean"),
            repeat_rate_pct=("repeat_customer_flag", lambda x: x.mean() * 100),
            delay_rate_pct=("delivery_delay_flag", lambda x: x.mean() * 100),
            avg_rating=("rating", "mean"),
        )
        .round(2)
        .reset_index()
        .sort_values("gross_profit", ascending=False)
    )

    discount = (
        df.groupby("discount_band", observed=True)
        .agg(
            orders=("order_id", "count"),
            avg_discount=("discount_amount", "mean"),
            avg_profit=("gross_profit", "mean"),
            repeat_rate_pct=("repeat_customer_flag", lambda x: x.mean() * 100),
            avg_rating=("rating", "mean"),
        )
        .round(2)
        .reset_index()
    )

    zone = (
        df.groupby(["city", "zone"])
        .agg(
            orders=("order_id", "count"),
            delay_rate_pct=("delivery_delay_flag", lambda x: x.mean() * 100),
            avg_delay_minutes=("delay_minutes", "mean"),
            avg_rating=("rating", "mean"),
            gross_profit=("gross_profit", "sum"),
        )
        .query("orders >= 30")
        .round(2)
        .reset_index()
        .sort_values(["delay_rate_pct", "avg_rating"], ascending=[False, True])
        .head(15)
    )

    cuisine = (
        df.groupby("cuisine_type")
        .agg(
            orders=("order_id", "count"),
            gross_profit=("gross_profit", "sum"),
            avg_margin_pct=("profit_margin_pct", "mean"),
            delay_rate_pct=("delivery_delay_flag", lambda x: x.mean() * 100),
        )
        .round(2)
        .reset_index()
        .sort_values("gross_profit", ascending=False)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    style_title(ws, "Food Delivery Profitability Dashboard")
    ws["A3"] = "KPI"
    ws["B3"] = "Value"
    ws["A3"].font = ws["B3"].font = Font(bold=True, color="FFFFFF")
    ws["A3"].fill = ws["B3"].fill = PatternFill("solid", fgColor=BLUE)
    for idx, row in enumerate(kpis.itertuples(index=False), 4):
        ws.cell(idx, 1, row.KPI)
        ws.cell(idx, 2, row.Value)
        ws.cell(idx, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE if idx % 2 == 0 else GREEN)
        ws.cell(idx, 2).fill = PatternFill("solid", fgColor=LIGHT_BLUE if idx % 2 == 0 else GREEN)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18

    ws["D3"] = "Business read"
    ws["D3"].font = Font(bold=True, color="FFFFFF")
    ws["D3"].fill = PatternFill("solid", fgColor=BLUE)
    reads = [
        "Order volume is healthy, but profit varies heavily by segment.",
        "Discounts should be measured against repeat rate and margin, not only orders.",
        "Delay hotspots need operational action before extra marketing spend.",
        "Restaurant-zone analysis is more useful than simple restaurant rankings.",
    ]
    for i, text in enumerate(reads, 4):
        ws.cell(i, 4, text)
    ws.column_dimensions["D"].width = 78

    chart_ws = wb.create_sheet("City Summary")
    write_table(chart_ws, city)
    chart = BarChart()
    chart.title = "Gross Profit by City"
    chart.y_axis.title = "Gross Profit"
    chart.x_axis.title = "City"
    data = Reference(chart_ws, min_col=3, min_row=1, max_row=len(city) + 1)
    cats = Reference(chart_ws, min_col=1, min_row=2, max_row=len(city) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    chart_ws.add_chart(chart, "I2")

    ws2 = wb.create_sheet("Discount Analysis")
    write_table(ws2, discount)
    chart2 = BarChart()
    chart2.title = "Repeat Rate by Discount Band"
    data2 = Reference(ws2, min_col=5, min_row=1, max_row=len(discount) + 1)
    cats2 = Reference(ws2, min_col=1, min_row=2, max_row=len(discount) + 1)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.height = 8
    chart2.width = 14
    ws2.add_chart(chart2, "H2")

    ws3 = wb.create_sheet("Zone Risk")
    write_table(ws3, zone)

    ws4 = wb.create_sheet("Cuisine Profit")
    write_table(ws4, cuisine)
    chart4 = BarChart()
    chart4.title = "Profit by Cuisine"
    data4 = Reference(ws4, min_col=3, min_row=1, max_row=len(cuisine) + 1)
    cats4 = Reference(ws4, min_col=1, min_row=2, max_row=len(cuisine) + 1)
    chart4.add_data(data4, titles_from_data=True)
    chart4.set_categories(cats4)
    chart4.height = 8
    chart4.width = 16
    ws4.add_chart(chart4, "H2")

    ws5 = wb.create_sheet("Cleaned Data Sample")
    sample_cols = [
        "order_id", "customer_id", "city", "zone", "cuisine_type",
        "delivery_time_minutes", "delivery_delay_flag", "order_value",
        "discount_amount", "gross_profit", "profit_margin_pct", "rating",
        "customer_tenure_group", "repeat_customer_flag",
    ]
    write_table(ws5, df[sample_cols].head(500))

    ws6 = wb.create_sheet("Notes")
    style_title(ws6, "Workbook Notes")
    notes = [
        "This workbook is a portfolio-ready Excel deliverable built from the cleaned dataset.",
        "Dashboard shows KPIs and business interpretation.",
        "City Summary, Discount Analysis, Zone Risk, and Cuisine Profit are designed for pivot-style review.",
        "Cleaned Data Sample keeps the file usable on GitHub while still showing row-level data.",
        "The full cleaned dataset remains available in data/processed/food_delivery_cleaned.csv.",
    ]
    for i, note in enumerate(notes, 3):
        ws6.cell(i, 1, note)
    ws6.column_dimensions["A"].width = 110

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Saved workbook: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()

